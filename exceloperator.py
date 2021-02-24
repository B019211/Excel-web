import openpyxl

def WriteExcel(fund_info_list):
    #rはエスケープシーケンスを無視
    wb = openpyxl.load_workbook('.\資産一覧.xlsx')
    ws = wb['ファンド']
    
    row = 2
    for fund in fund_info_list:
        col = 1
        #6,7列目は更新対象外
        ws.cell(column=col, row=row, value=fund.name)
        col += 1
        ws.cell(column=col, row=row, value=fund.company)
        col += 1
        ws.cell(column=col, row=row, value=fund.category)
        col += 1
        ws.cell(column=col, row=row, value=fund.baseprice)
        col += 1
        ws.cell(column=col, row=row, value=float(fund.assets.replace(',', '')))
        col += 3
        ws.cell(column=col, row=row, value=int(fund.allotment))
        col += 1
        if fund.commision == 'なし':
            ws.cell(column=col, row=row, value=0)
        else:
            ws.cell(column=col, row=row, value=fund.commision)
        col += 1
        ws.cell(column=col, row=row, value=fund.cost)
        row += 1
        
    wb.save('.\資産一覧.xlsx')
        

        